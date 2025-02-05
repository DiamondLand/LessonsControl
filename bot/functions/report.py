import openpyxl
import os

from datetime import date
from aiogram import Bot
from aiogram.types import FSInputFile
from openpyxl.styles import Font, PatternFill, Border, Side

from database.services import get_attendance_for_today
from config.cfg import ADMIN_CHATS


async def generate_attendance_report(bot: Bot):
    """
    Генерирует отчёт о присутствии пользователей на занятиях за сегодня и сохраняет его в Excel в папку assets.

    Args:
        bot (Bot): Объект бота для отправки файла.
    """
    today = date.today()

    # Получаем список пользователей, которые были на занятиях сегодня
    attendance_data = await get_attendance_for_today(today=today)

    # Подготовим данные для Excel: Имя и Фамилия, Группа
    user_data = []
    for attendance in attendance_data:
        user = attendance.user
        user_data.append({
            'fullname': f"{user.firstname} {user.lastname}",
            'group': user.group
        })

    # Сортируем по фамилии и имени
    user_data.sort(key=lambda x: (x['fullname']))

    # Создаём новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Посещаемость {today.strftime('%d.%m.%Y')}"

    # Заголовки
    ws.append(["Имя Фамилия, группа"])
    header_cell = ws["A1"]
    header_cell.font = Font(bold=True, size=14, color="FF0000")
    header_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_cell.border = Border(bottom=Side(style='thin', color="000000"))
    header_cell.number_format = '#,##0'

    # Добавляем данные
    for entry in user_data:
        ws.append([f"{entry['fullname']}, {entry['group']}"])

    # Форматируем данные
    data_font = Font(size=12)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = data_font

    # Форматируем ширину столбца для лучшего отображения
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 25  # Увеличиваем ширину столбца

    # Сохраняем файл в папку assets
    file_path = os.path.join(bot.ASSETS_PATH, f"{today.strftime('%d%m%Y')}.xlsx")
    wb.save(file_path)

    # Отправляем файл пользователю
    for chat_id in ADMIN_CHATS:
        await bot.send_document(
            chat_id=chat_id,
            caption=f"Посещаемость за {today.strftime('%d.%m.%Y')}",
            document=FSInputFile(path=file_path, filename=f"{today.strftime('%d%m%Y')}.xlsx")
        )
